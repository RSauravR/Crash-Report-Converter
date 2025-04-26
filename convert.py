import pdfplumber
import pandas as pd
from tkinter import Tk, filedialog
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def select_pdf():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    return file_path

def extract_vehicle_info(page_text):
    fields = {
        "Test Protocol": r"Test Protocol:\s*(.*)",
        "Vehicle ID": r"Vehicle ID:\s*(.*)",
        "Vehicle Type": r"Vehicle Type:\s*(.*)",
        "Doors": r"Doors:\s*(.*)",
        "Weight (Dry)": r"Weight \(Dry\):\s*(.*)",
        "Weight (Wet)": r"Weight \(Wet\):\s*(.*)",
        "Powertrain": r"Powertrain:\s*(.*)",
        "Test Speed": r"Test Speed:\s*(.*)",
        "Barrier Type": r"Barrier Type:\s*(.*)",
        "Impact Type": r"Impact Type:\s*(.*)"
    }
    data = {}
    for key, pattern in fields.items():
        match = re.search(pattern, page_text)
        data[key] = match.group(1).strip() if match else "N/A"
    return data

def extract_injury_data(page_text):
    injuries = {}
    lines = page_text.splitlines()
    for line in lines:
        match = re.match(r"(.+):\s*(\d+)%", line)
        if match:
            metric = match.group(1).strip()
            value = int(match.group(2))
            injuries[metric] = value
    return injuries

def process_pdf_to_excel(pdf_path, output_path):
    with pdfplumber.open(pdf_path) as pdf:
        vehicle_data = extract_vehicle_info(pdf.pages[1].extract_text())
        injury_data = {}
        for i in range(2, len(pdf.pages)):
            page_text = pdf.pages[i].extract_text()
            injury_data.update(extract_injury_data(page_text))

    vehicle_df = pd.DataFrame(vehicle_data.items(), columns=["Attribute", "Value"])
    injury_df = pd.DataFrame(injury_data.items(), columns=["Body Part", "Injury %"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        vehicle_df.to_excel(writer, sheet_name="Vehicle Info", index=False)
        injury_df.to_excel(writer, sheet_name="Injuries", index=False)

    wb = load_workbook(output_path)
    ws = wb["Injuries"]
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if isinstance(cell.value, int):
                if cell.value <= 80:
                    cell.fill = yellow
                elif cell.value > 100:
                    cell.fill = red
    wb.save(output_path)
    print(f"✅ Excel file saved to: {output_path}")

if __name__ == "__main__":
    pdf_file = select_pdf()
    if pdf_file:
        output_file = pdf_file.replace(".pdf", "_extracted.xlsx")
        process_pdf_to_excel(pdf_file, output_file)
    else:
        print("❌ No PDF selected.")
